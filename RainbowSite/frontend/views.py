from django.shortcuts import render, redirect
from .models import blog, result
from django.contrib import messages
import openpyxl
from openpyxl import load_workbook
import os
# Create your views here.


def home(request):
    return render(request, 'frontend/home.html')

def user_home(request):
    if request.method == 'POST':
        if 'results' in request.POST:
            f1 = request.FILES['form1']
            f2 = request.FILES['form2']
            f3 = request.FILES['form3']
            f4 = request.FILES['form4']
            f5 = request.FILES['form5']
            lsa = request.FILES['lsa']
            lss = request.FILES['lss']
            uss = request.FILES['uss']
            usa = request.FILES['usa']
            try:
                res = results.objects.first()
                modif = res(
                    f1=f1,
                    f2=f2,
                    f3=f3,
                    f4=f4,
                    f5=f5,
                    lss=lss,
                    lsa=lsa,
                    usl=uss,
                    usa=usa,
                )
                modif.save()
                messages.success(request,'results updated')
                return redirect('user_home')
            except:
                res = result.objects.create(
                    f1=f1,
                    f2=f2,
                    f3=f3,
                    f4=f4,
                    f5=f5,
                    lss=lss,
                    lsa=lsa,
                    usl=uss,
                    usa=usa,
                )
                res.save()
                messages.success(request, 'results updated')
                return redirect('user_home')
        else:
            title = request.POST['title']
            description = request.POST['description']
            post_image = request.FILES['post_image']
            _blog = blog.objects.create(author=request.user.secretary, title = title, description = description, image = post_image)
            _blog.save()
            print('saved')
            messages.success(request, "Blog post added successfully")
            return redirect('user_home')
    else:
        blogs = blog.objects.all()
        rlt = result.objects.first()

    context = {
        'blogs' : blogs,
        'rlt':rlt,
    }
    return render(request, 'backend/user.html', context)


def gallery(request):
    blogs = blog.objects.all()

    context = {
        'blogs' : blogs
    }
    return render(request, 'frontend/gallery.html', context)




def my_blog(request):
    blogs = blog.objects.all()

    context = {
        'blogs' : blogs
    }
    return render(request, 'frontend/blog.html',context)

def delete(request, id):
    blog.objects.get(pk=id).delete()
    return redirect('user_home')


def activateView(request):
    rlt = result.objects.first()
    rlt.active = True
    rlt.save()
    return redirect('user_home')
def deactivateView(request):
    rlt = result.objects.first()
    rlt.active = False
    rlt.save()
    return redirect('user_home')


def results(request):
    if request.method == 'POST':
        result_=result.objects.first()
        if result.active == True:
            mark=list()
            dic = {'name'}
            item = list()
            mat = request.POST['matricule']
            class_= request.POST['class']
            sequence= request.POST['sequence']

            file_path = 'media/results/' + class_ + '.xlsx'
            #file_path = 'media/results/form1.xlsx'
            #file_path = os.path.join(os.path.dirname(__file__), 'form1.xlsx')
            workbook = load_workbook(filename=file_path, keep_vba=True, data_only=True)
            sheet = workbook[sequence]
            for col in sheet.iter_cols(max_col=1, min_col=1, min_row=1):
                for cell in col:
                    if cell.value == int(mat):
                        for row in sheet.iter_rows(max_col=23, min_row=cell.row, max_row=cell.row):
                            for cel in row:
                                print(cel.row)
                                mark.append(cel.value)
            for col2 in sheet.iter_rows(max_col=23, min_row=10, max_row=10):
                for cells in col2:
                    item.append(cells.value)
            zipped=zip(item, mark)
            
            context = {
                'marks':mark,
                'items':item,
                'zipped': zipped,
            }
            workbook.close()
            return render(request, 'frontend/report.html', context)
        else:
            return render(request, 'frontend/noresults.html')
        
    return render(request, 'frontend/result.html')


def deleteResult(request):
    result.objects.first().delete()
    return redirect('user_home')